"""
Excel Bank Statement Parser — v2 (Production-Ready)
────────────────────────────────────────────────────
Strict, rule-based parser for Saudi bank statements in Excel format.

v2 enhancements:
  ✓ Multi-sheet / multi-account support
  ✓ Internal transfer detection & exclusion
  ✓ Reversal transaction pairing
  ✓ Smarter duplicate detection (row-number aware)
  ✓ Debit/credit sign normalisation
  ✓ Amount-only column → sign-based debit/credit split
  ✓ Arabic numeral conversion
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Any

import openpyxl
from loguru import logger

# ─── Column aliases ────────────────────────────────────

COLUMN_MAP: dict[str, list[str]] = {
    "date": [
        "date", "تاريخ", "التاريخ", "transaction date", "تاريخ العملية",
        "value date", "تاريخ القيمة", "posting date",
    ],
    "description": [
        "description", "الوصف", "البيان", "details", "التفاصيل",
        "narration", "particulars", "transaction description",
        "وصف العملية", "بيان",
    ],
    "debit": [
        "debit", "مدين", "المدين", "withdrawal", "سحب", "مسحوب",
        "debit amount", "مبلغ مدين", "withdrawals",
    ],
    "credit": [
        "credit", "دائن", "الدائن", "deposit", "إيداع", "مودع",
        "credit amount", "مبلغ دائن", "deposits",
    ],
    "balance": [
        "balance", "الرصيد", "رصيد", "running balance",
        "الرصيد المتاح", "closing balance",
    ],
    "amount": [
        "amount", "المبلغ", "مبلغ", "value", "القيمة",
    ],
    "reference": [
        "reference", "المرجع", "رقم المرجع", "ref", "رقم العملية",
        "transaction id", "رقم الحركة",
    ],
    "account": [
        "account", "الحساب", "رقم الحساب", "account number",
        "account no", "iban",
    ],
}

# ─── Internal transfer patterns ────────────────────────

INTERNAL_TRANSFER_PATTERNS: list[str] = [
    "تحويل بين الحسابات", "تحويل داخلي", "تحويل من حساب",
    "تحويل إلى حساب", "internal transfer", "between accounts",
    "own account transfer", "standing order own",
    "self transfer", "تحويل ذاتي", "من حسابي",
    "إلى حسابي", "حركة داخلية",
]

# ─── Reversal patterns ─────────────────────────────────

REVERSAL_PATTERNS: list[str] = [
    "reversal", "عكسية", "عملية عكسية", "عكس",
    "تصحيح", "correction", "reversed", "إلغاء",
    "cancelled", "storno", "void",
    "مرتجع", "استرداد", "refund",
]


@dataclass
class Transaction:
    row_number: int
    sheet_name: str
    date: date | None
    description: str
    debit: float
    credit: float
    balance: float | None
    reference: str = ""
    account: str = ""
    raw_values: dict[str, Any] = field(default_factory=dict)
    category: str = ""
    is_duplicate: bool = False
    is_outlier: bool = False
    is_internal_transfer: bool = False
    is_reversal: bool = False
    reversal_pair_row: int | None = None


@dataclass
class AccountSummary:
    sheet_name: str
    account_id: str
    transaction_count: int = 0
    total_credit: float = 0.0
    total_debit: float = 0.0


@dataclass
class ParseResult:
    transactions: list[Transaction]
    total_rows: int
    parsed_rows: int
    skipped_rows: int
    duplicate_count: int
    outlier_count: int
    internal_transfer_count: int
    reversal_count: int
    date_range_start: date | None
    date_range_end: date | None
    months_covered: int
    columns_found: list[str]
    columns_missing: list[str]
    issues: list[str]
    confidence: float
    accounts: list[AccountSummary] = field(default_factory=list)
    sheets_parsed: int = 0


# ─── Helpers ───────────────────────────────────────────

def _normalize(text: str) -> str:
    return re.sub(r"\s+", " ", str(text).strip().lower())


def _match_column(header: str) -> str | None:
    norm = _normalize(header)
    for canonical, aliases in COLUMN_MAP.items():
        for alias in aliases:
            if alias in norm or norm in alias:
                return canonical
    return None


def _parse_number(val: Any) -> float:
    if val is None or val == "" or val == "-":
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    s = s.replace(",", "").replace("،", "").replace(" ", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    is_negative = False
    if s.upper().endswith("CR"):
        s = s[:-2].strip()
    elif s.upper().endswith("DR"):
        s = s[:-2].strip()
        is_negative = True
    arabic_map = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
    s = s.translate(arabic_map)
    try:
        result = float(s)
        return -result if is_negative else result
    except ValueError:
        return 0.0


def _parse_date(val: Any) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in (
        "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d",
        "%d-%m-%Y", "%d.%m.%Y", "%Y%m%d",
        "%d %b %Y", "%d %B %Y",
    ):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _count_months(start: date, end: date) -> int:
    if start > end:
        start, end = end, start
    months: set[tuple[int, int]] = set()
    current = start
    while current <= end:
        months.add((current.year, current.month))
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1, day=1)
        else:
            current = current.replace(month=current.month + 1, day=1)
    return len(months)


def _is_internal_transfer(desc: str) -> bool:
    desc_lower = desc.lower().strip()
    return any(p in desc_lower for p in INTERNAL_TRANSFER_PATTERNS)


def _is_reversal(desc: str) -> bool:
    desc_lower = desc.lower().strip()
    return any(p in desc_lower for p in REVERSAL_PATTERNS)


# ─── Duplicate detection ──────────────────────────────

def _flag_duplicates(txns: list[Transaction]) -> int:
    seen: dict[tuple, int] = {}
    count = 0
    for idx, t in enumerate(txns):
        key = (t.date, _normalize(t.description), round(t.debit, 2), round(t.credit, 2), t.sheet_name)
        if key in seen:
            t.is_duplicate = True
            count += 1
        else:
            seen[key] = idx
    return count


# ─── Reversal pairing ────────────────────────────────

def _pair_reversals(txns: list[Transaction]) -> int:
    count = 0
    reversal_indices = [i for i, t in enumerate(txns) if t.is_reversal]
    for rev_idx in reversal_indices:
        rev = txns[rev_idx]
        for orig_idx in range(rev_idx - 1, -1, -1):
            orig = txns[orig_idx]
            if orig.is_reversal or orig.reversal_pair_row is not None:
                continue
            date_close = (
                orig.date == rev.date or
                (orig.date and rev.date and abs((rev.date - orig.date).days) <= 3)
            )
            if date_close:
                if (
                    (round(orig.debit, 2) == round(rev.credit, 2) and orig.debit > 0) or
                    (round(orig.credit, 2) == round(rev.debit, 2) and orig.credit > 0)
                ):
                    rev.reversal_pair_row = orig.row_number
                    orig.reversal_pair_row = rev.row_number
                    count += 1
                    break
            if rev.reference and orig.reference and rev.reference == orig.reference:
                rev.reversal_pair_row = orig.row_number
                orig.reversal_pair_row = rev.row_number
                count += 1
                break
    return count


# ─── Outlier detection ────────────────────────────────

def _flag_outliers(txns: list[Transaction], factor: float = 3.0) -> int:
    amounts = []
    for t in txns:
        if t.is_duplicate or t.is_internal_transfer:
            continue
        amt = t.credit if t.credit > 0 else t.debit
        if amt > 0:
            amounts.append(amt)
    if len(amounts) < 10:
        return 0
    amounts.sort()
    q1 = amounts[len(amounts) // 4]
    q3 = amounts[3 * len(amounts) // 4]
    iqr = q3 - q1
    upper = q3 + factor * iqr
    lower = q1 - factor * iqr
    count = 0
    for t in txns:
        if t.is_duplicate or t.is_internal_transfer:
            continue
        amt = t.credit if t.credit > 0 else t.debit
        if amt > 0 and (amt < lower or amt > upper):
            t.is_outlier = True
            count += 1
    return count


# ─── Parse single sheet ──────────────────────────────

def _parse_sheet(ws: Any, sheet_name: str, file_bytes: bytes) -> tuple[list[Transaction], list[str], list[str], list[str], int]:
    issues: list[str] = []
    header_row_idx: int | None = None
    col_mapping: dict[str, int] = {}

    for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
        if row_idx > 25:
            break
        candidate: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            val = cell.value
            if val is None:
                continue
            canonical = _match_column(str(val))
            if canonical and canonical not in candidate:
                candidate[canonical] = col_idx
        if "date" in candidate and ("debit" in candidate or "credit" in candidate or "amount" in candidate):
            header_row_idx = row_idx
            col_mapping = candidate
            break

    if header_row_idx is None:
        return [], [], ["date", "debit", "credit"], [f"[{sheet_name}] لم يتم العثور على صف العناوين"], 0

    columns_found = list(col_mapping.keys())
    columns_missing = [c for c in ["date", "description"] if c not in col_mapping]
    has_split = "debit" in col_mapping and "credit" in col_mapping
    has_amount = "amount" in col_mapping and not has_split

    wb2 = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    ws2 = wb2[sheet_name] if sheet_name in wb2.sheetnames else wb2.active

    transactions: list[Transaction] = []
    skipped = 0

    for row_idx, row in enumerate(ws2.iter_rows(values_only=True), start=1):
        if row_idx <= header_row_idx:
            continue
        raw: dict[str, Any] = {}
        for canonical, col_idx in col_mapping.items():
            raw[canonical] = row[col_idx] if col_idx < len(row) else None
        if all(v is None or v == "" for v in raw.values()):
            skipped += 1
            continue

        dt = _parse_date(raw.get("date"))
        desc = str(raw.get("description", "") or "").strip()
        ref = str(raw.get("reference", "") or "").strip()
        acct = str(raw.get("account", "") or "").strip()

        if has_split:
            debit = abs(_parse_number(raw.get("debit")))
            credit = abs(_parse_number(raw.get("credit")))
        elif has_amount:
            raw_amt = _parse_number(raw.get("amount"))
            credit = raw_amt if raw_amt >= 0 else 0.0
            debit = abs(raw_amt) if raw_amt < 0 else 0.0
        else:
            debit = abs(_parse_number(raw.get("debit", 0)))
            credit = abs(_parse_number(raw.get("credit", 0)))

        balance = _parse_number(raw.get("balance")) if "balance" in raw else None

        if dt is None and debit == 0 and credit == 0:
            skipped += 1
            continue

        transactions.append(Transaction(
            row_number=row_idx,
            sheet_name=sheet_name,
            date=dt,
            description=desc,
            debit=debit,
            credit=credit,
            balance=balance,
            reference=ref,
            account=acct,
            raw_values=raw,
            is_internal_transfer=_is_internal_transfer(desc),
            is_reversal=_is_reversal(desc),
        ))

    wb2.close()
    return transactions, columns_found, columns_missing, issues, skipped


# ─── Main parser ──────────────────────────────────────

def parse_bank_statement(file_bytes: bytes, required_months: int = 6) -> ParseResult:
    issues: list[str] = []
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        logger.error(f"Failed to open workbook: {e}")
        return ParseResult(
            transactions=[], total_rows=0, parsed_rows=0, skipped_rows=0,
            duplicate_count=0, outlier_count=0,
            internal_transfer_count=0, reversal_count=0,
            date_range_start=None, date_range_end=None, months_covered=0,
            columns_found=[], columns_missing=["date", "description", "debit", "credit"],
            issues=[f"فشل في فتح الملف: {e}"], confidence=0.0,
        )

    all_txns: list[Transaction] = []
    all_cols_found: list[str] = []
    all_cols_missing: list[str] = []
    total_skipped = 0
    accounts: list[AccountSummary] = []
    sheets_parsed = 0

    for sname in wb.sheetnames:
        ws = wb[sname]
        txns, cf, cm, si, sk = _parse_sheet(ws, sname, file_bytes)
        if txns:
            sheets_parsed += 1
            all_txns.extend(txns)
            for c in cf:
                if c not in all_cols_found:
                    all_cols_found.append(c)
            acct_id = txns[0].account if txns[0].account else sname
            accounts.append(AccountSummary(
                sheet_name=sname, account_id=acct_id,
                transaction_count=len(txns),
                total_credit=sum(t.credit for t in txns),
                total_debit=sum(t.debit for t in txns),
            ))
        total_skipped += sk
        issues.extend(si)
        all_cols_missing = cm

    wb.close()

    total_rows = len(all_txns) + total_skipped
    parsed_rows = len(all_txns)

    if parsed_rows == 0:
        return ParseResult(
            transactions=[], total_rows=total_rows, parsed_rows=0, skipped_rows=total_skipped,
            duplicate_count=0, outlier_count=0,
            internal_transfer_count=0, reversal_count=0,
            date_range_start=None, date_range_end=None, months_covered=0,
            columns_found=all_cols_found, columns_missing=all_cols_missing,
            issues=issues + ["لم يتم العثور على أي معاملات صالحة"],
            confidence=0.0, accounts=accounts, sheets_parsed=sheets_parsed,
        )

    dates = [t.date for t in all_txns if t.date is not None]
    date_start = min(dates) if dates else None
    date_end = max(dates) if dates else None
    months_covered = _count_months(date_start, date_end) if dates else 0
    if not dates:
        issues.append("لم يتم العثور على تواريخ صالحة")
    if months_covered < required_months:
        issues.append(f"التغطية الزمنية ({months_covered} شهر) أقل من المطلوب ({required_months} شهر)")

    dup_count = _flag_duplicates(all_txns)
    if dup_count > 0:
        issues.append(f"تم اكتشاف {dup_count} معاملة مكررة (لن تُحتسب)")

    internal_count = sum(1 for t in all_txns if t.is_internal_transfer)
    if internal_count > 0:
        issues.append(f"تم اكتشاف {internal_count} تحويل داخلي (مُستبعد)")

    reversal_count = _pair_reversals(all_txns)
    if reversal_count > 0:
        issues.append(f"تم ربط {reversal_count} عملية عكسية")

    outlier_count = _flag_outliers(all_txns)
    if outlier_count > 0:
        issues.append(f"تم اكتشاف {outlier_count} معاملة غير اعتيادية")

    confidence = _compute_confidence(
        parsed_rows=parsed_rows, total_rows=total_rows, skipped=total_skipped,
        months_covered=months_covered, required_months=required_months,
        columns_missing=all_cols_missing, dup_count=dup_count,
        outlier_count=outlier_count, has_dates=len(dates) > 0, internal_count=internal_count,
    )

    logger.info(
        f"Parsed {parsed_rows}/{total_rows} rows across {sheets_parsed} sheet(s), "
        f"{months_covered}mo, {dup_count} dups, {internal_count} internal, "
        f"{reversal_count} reversals, confidence={confidence:.2f}"
    )

    return ParseResult(
        transactions=all_txns, total_rows=total_rows,
        parsed_rows=parsed_rows, skipped_rows=total_skipped,
        duplicate_count=dup_count, outlier_count=outlier_count,
        internal_transfer_count=internal_count, reversal_count=reversal_count,
        date_range_start=date_start, date_range_end=date_end,
        months_covered=months_covered,
        columns_found=all_cols_found, columns_missing=all_cols_missing,
        issues=issues, confidence=confidence,
        accounts=accounts, sheets_parsed=sheets_parsed,
    )


def _compute_confidence(
    parsed_rows: int, total_rows: int, skipped: int,
    months_covered: int, required_months: int,
    columns_missing: list[str], dup_count: int,
    outlier_count: int, has_dates: bool, internal_count: int = 0,
) -> float:
    score = 0.0
    required = ["date", "description", "debit", "credit"]
    found = len(required) - len([c for c in columns_missing if c in required])
    score += 0.25 * (found / len(required))
    if required_months > 0 and has_dates:
        score += 0.25 * min(months_covered / required_months, 1.0)
    elif has_dates:
        score += 0.20
    if total_rows > 0:
        score += 0.20 * (parsed_rows / total_rows)
    if parsed_rows > 0:
        score += 0.15 * max(0, 1.0 - (dup_count / parsed_rows) * 5)
        score += 0.15 * max(0, 1.0 - ((outlier_count + internal_count * 0.1) / parsed_rows) * 3)
    return round(max(0.0, min(1.0, score)), 4)
